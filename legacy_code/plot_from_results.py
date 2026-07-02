# -*- coding: utf-8 -*-
from __future__ import print_function, division
import csv, os
import numpy as np

# 服务器无显示环境也能保存图
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

model_analyse="wrn_cifar100"

if model_analyse =="cvt":
    INPUT_FILES = [
        "/home/ywang/Article/vgg_gradmax/VTs-Drloc-master/12grows_epoch200_adj/run1/results_run_from_grow.txt",
        "/home/ywang/Article/vgg_gradmax/VTs-Drloc-master/12grows_epoch200_adj/run2/results_run_from_grow.txt",
        "/home/ywang/Article/vgg_gradmax/VTs-Drloc-master/12grows_epoch200_adj/run3/results_run_from_grow.txt"

    ]

    OUTPUT_DIR = "/home/ywang/Article/vgg_gradmax/VTs-Drloc-master/12grows_epoch200_adj"
elif model_analyse =="vgg_cifar10":
    INPUT_FILES = [
        "/home/ywang/Article/vgg_gradmax/outputs_compare/vgg11/cifar10/20260109_1740_seed42_lr0.05/results_run_1.txt",
        "/home/ywang/Article/vgg_gradmax/outputs_compare/vgg11/cifar10/20260109_2045_seed17_lr0.05/results_run_1.txt",
        "/home/ywang/Article/vgg_gradmax/outputs_compare/vgg11/cifar10/20260110_105252_run01_seed1501552845_lr0.05/results_run_1.txt"

    ]

    OUTPUT_DIR = "/home/ywang/Article/vgg_gradmax/outputs_compare/vgg11/cifar10"
elif model_analyse =="vgg_cifar100":
    INPUT_FILES = [
        "/home/ywang/Article/vgg_gradmax/outputs_compare/vgg11/cifar100/20260110_004313_run01_seed17_lr0.05/results_run_1.txt",
        "/home/ywang/Article/vgg_gradmax/outputs_compare/vgg11/cifar100/20260110_035601_run02_seed17_lr0.05/results_run_2.txt",
        "/home/ywang/Article/vgg_gradmax/outputs_compare/vgg11/cifar100/20260110_065453_run03_seed17_lr0.05/results_run_3.txt"

    ]

    OUTPUT_DIR = "/home/ywang/Article/vgg_gradmax/outputs_compare/vgg11/cifar100"
elif model_analyse =="wrn_cifar10":
    INPUT_FILES = [
        "/home/ywang/Article/vgg_gradmax/outputs_compare/wrn/cifar10/rlt/results_run_1.txt",
        "/home/ywang/Article/vgg_gradmax/outputs_compare/wrn/cifar10/rlt/results_run_2.txt",
        "/home/ywang/Article/vgg_gradmax/outputs_compare/wrn/cifar10/rlt/results_run_3.txt"

    ]
    OUTPUT_DIR = "/home/ywang/Article/vgg_gradmax/outputs_compare/wrn/cifar10/rlt"
    
elif model_analyse =="wrn_cifar100_lr01":
    INPUT_FILES = [
        "/home/ywang/Article/vgg_gradmax/outputs_compare/wrn/cifar100/lr01/20260114_165524_run01_seed1501552845_lr0.1/results_run_1.txt",
        "/home/ywang/Article/vgg_gradmax/outputs_compare/wrn/cifar100/lr01/20260114_211027_run02_seed488200390_lr0.1/results_run_2.txt",
        "/home/ywang/Article/vgg_gradmax/outputs_compare/wrn/cifar100/lr01/20260115_012452_run03_seed1693606510_lr0.1/results_run_3.txt"

    ]

    OUTPUT_DIR = "/home/ywang/Article/vgg_gradmax/outputs_compare/wrn/cifar100/lr01"   

  
elif model_analyse =="wrn_cifar100":
    INPUT_FILES = [
        "/home/ywang/Article/vgg_gradmax/outputs_compare/wrn/cifar100/lr001/20260127_162745_run01_seed2_lr0.01/results_run_1.txt",
        "/home/ywang/Article/vgg_gradmax/outputs_compare/wrn/cifar100/lr001/20260127_195509_run02_seed23_lr0.01/results_run_2.txt",
        "/home/ywang/Article/vgg_gradmax/outputs_compare/wrn/cifar100/lr001/20260227_193342_run01_seed22_lr0.01/results_run_1.txt"

    ]

    OUTPUT_DIR = "/home/ywang/Article/vgg_gradmax/outputs_compare/wrn/cifar100/rlt"
    
elif model_analyse =="vit_cifar10":
    INPUT_FILES = [
        "/home/ywang/Article/vgg_gradmax/outputs_compare/vit/cifar10/new_tech/layers8_heads4_hiddim256_drop0_masterSeed145/run01_seed145/results_run_1.txt",
        "/home/ywang/Article/vgg_gradmax/outputs_compare/vit/cifar10/new_tech/layers8_heads4_hiddim256_drop0_masterSeed145/run02_seed146/results_run_2.txt",
        "/home/ywang/Article/vgg_gradmax/outputs_compare/vit/cifar10/new_tech/layers8_heads4_hiddim256_drop0_masterSeed145/run03_seed147/results_run_3.txt"

    ]

    OUTPUT_DIR = "/home/ywang/Article/vgg_gradmax/outputs_compare/wrn/cifar10/rlt"   
    
elif model_analyse =="mlp":
    INPUT_FILES = [
        "/home/ywang/Article/vgg_gradmax/outputs_compare/mlp/20260110_111830_run01_seed1501552845_lr0.05/results_run_1.txt",
        "/home/ywang/Article/vgg_gradmax/outputs_compare/mlp/20260110_114152_run02_seed488200390_lr0.05/results_run_2.txt",
        "/home/ywang/Article/vgg_gradmax/outputs_compare/mlp/20260110_120611_run03_seed1693606510_lr0.05/results_run_3.txt"

    ]

    OUTPUT_DIR = "/home/ywang/Article/vgg_gradmax/outputs_compare/mlp"   



    
# 只画这些模型（按需改）；如果你想自动画文件里出现的全部模型，把它设为 None
MODELS_TO_PLOT = ["c", "a", "b", "d", "e", "f"]

# True：只用三个 run 都存在的 epoch（推荐，最稳）
USE_EPOCH_INTERSECTION = True

FIG_TITLE = "Test Accuracy vs Time (mean ± std over 3 runs)"


# ===== 新增：统计“达到 Big Net(c) 同等 TestAcc 所需时间” =====
# 解释：
#   - 先在每个 run 里计算 Big Net(c) 的目标 TestAcc（默认：最后5个有效 epoch 的平均 = final5）
#   - 再在同一个 run 里，对 a/b/d/e/f 找到最早满足 TestAcc >= 目标值 的时间点
#   - 可选：要求连续 MATCH_STREAK 个 epoch 都满足（更稳健，抗抖动）
MATCH_REF_MODEL = "c"
MATCH_COMPARE_MODELS = ["a", "b", "d", "e", "f"]   # 只统计这些；你也可以设为 None 表示“除 c 外所有模型”
MATCH_REF_ACC_MODE = "final5"  # "final5" | "last" | "best"
MATCH_REF_LAST_K = 5
MATCH_STREAK = 3               # 连续多少个 epoch 满足 >= target 才算“达到”
MATCH_USE_END_OF_STREAK = True # True：取窗口最后一个 epoch 的时间（更保守）；False：取窗口第一个
MATCH_ACC_EPS = 0.0            # 容差：判断 acc >= target - eps


# 新增三张图（与 TestAcc vs Time 风格一致）：
#   1) TrainLoss vs Time
#   2) TrainAcc  vs Time
#   3) TestLoss  vs Time


def _safe_float(x, default=np.nan):
    try:
        return float(x)
    except Exception:
        return default


def _safe_int(x, default=None):
    try:
        return int(x)
    except Exception:
        return default


def read_results_tsv(path):
    """读取单个 results_run_*.txt，返回 dict: model -> {epoch -> (time, test_acc)}"""
    data = {}
    with open(path, "r") as f:
        reader = csv.DictReader(f, delimiter="\t")
        required = ["Epoch", "Model", "TestAcc", "Time(s)"]
        for k in required:
            if k not in reader.fieldnames:
                raise ValueError(
                    "文件缺少列: %s (实际列: %s) -> %s"
                    % (k, reader.fieldnames, path)
                )

        for row in reader:
            epoch = _safe_int(row.get("Epoch"))
            model = row.get("Model")
            test_acc = _safe_float(row.get("TestAcc"))
            t = _safe_float(row.get("Time(s)"))

            if epoch is None or model is None:
                continue

            if model not in data:
                data[model] = {}
            # 同一个 (model, epoch) 多次出现则以最后一次为准
            data[model][epoch] = (t, test_acc)

    return data


def read_results_tsv_allmetrics(path):
    """读取单个 results_run_*.txt，返回 dict:

    model -> {epoch -> {
        'time': float,
        'train_loss': float,
        'train_acc': float,
        'test_loss': float,
        'test_acc': float,
    }}

    说明：
      - 只强制要求存在 Epoch / Model / Time(s)
      - TrainLoss/TrainAcc/TestLoss/TestAcc 若缺列则会被填成 nan
    """
    data = {}
    with open(path, "r") as f:
        reader = csv.DictReader(f, delimiter="\t")
        required = ["Epoch", "Model", "Time(s)"]
        for k in required:
            if k not in reader.fieldnames:
                raise ValueError(
                    "文件缺少列: %s (实际列: %s) -> %s" % (k, reader.fieldnames, path)
                )

        for row in reader:
            epoch = _safe_int(row.get("Epoch"))
            model = row.get("Model")
            t = _safe_float(row.get("Time(s)"))

            if epoch is None or model is None:
                continue

            rec = {
                "time": t,
                "train_loss": _safe_float(row.get("TrainLoss")),
                "train_acc": _safe_float(row.get("TrainAcc")),
                "test_loss": _safe_float(row.get("TestLoss")),
                "test_acc": _safe_float(row.get("TestAcc")),
            }

            if model not in data:
                data[model] = {}
            # 同一个 (model, epoch) 多次出现则以最后一次为准
            data[model][epoch] = rec

    return data


def aggregate_runs(run_dicts, models=None, use_epoch_intersection=True):
    """聚合多次 run，按 (Model, Epoch) 对齐，计算 mean/std。"""
    all_models = set()
    for rd in run_dicts:
        all_models.update(rd.keys())

    if models is not None:
        use_models = [m for m in models if m in all_models]
    else:
        use_models = sorted(list(all_models))

    agg = {}
    for m in use_models:
        epoch_sets = []
        for rd in run_dicts:
            epoch_sets.append(set(rd.get(m, {}).keys()))

        if use_epoch_intersection:
            common = None
            for s in epoch_sets:
                common = s if common is None else common.intersection(s)
            epochs = sorted(list(common)) if common is not None else []
        else:
            union = set()
            for s in epoch_sets:
                union.update(s)
            epochs = sorted(list(union))

        if len(epochs) == 0:
            continue

        mean_time, mean_acc, std_acc, n_list = [], [], [], []

        for ep in epochs:
            times, accs = [], []
            for rd in run_dicts:
                if m in rd and ep in rd[m]:
                    t, a = rd[m][ep]
                    if np.isfinite(t):
                        times.append(t)
                    if np.isfinite(a):
                        accs.append(a)

            n = len(accs)
            n_list.append(n)

            mean_time.append(float(np.mean(times)) if len(times) else np.nan)

            if len(accs):
                mean_acc.append(float(np.mean(accs)))
                std_acc.append(float(np.std(accs, ddof=1)) if len(accs) > 1 else 0.0)
            else:
                mean_acc.append(np.nan)
                std_acc.append(np.nan)

        agg[m] = {
            "epochs": np.array(epochs, dtype=np.int64),
            "mean_time": np.array(mean_time, dtype=np.float64),
            "mean_acc": np.array(mean_acc, dtype=np.float64),
            "std_acc": np.array(std_acc, dtype=np.float64),
            "n": np.array(n_list, dtype=np.int64),
        }

    return agg


def aggregate_runs_metric(run_dicts, value_field, models=None, use_epoch_intersection=True):
    """聚合多次 run（支持任意指标），按 (Model, Epoch) 对齐，计算 mean/std。

    参数：
      - run_dicts: list[dict]，每个元素为 read_results_tsv_allmetrics 的返回值
      - value_field: 'train_loss' / 'train_acc' / 'test_loss' / 'test_acc' 等
    返回：
      agg[model] = {
        'epochs': np.array,
        'mean_time': np.array,
        'mean_value': np.array,
        'std_value': np.array,
        'n': np.array,
      }
    """
    all_models = set()
    for rd in run_dicts:
        all_models.update(rd.keys())

    if models is not None:
        use_models = [m for m in models if m in all_models]
    else:
        use_models = sorted(list(all_models))

    agg = {}
    for m in use_models:
        epoch_sets = []
        for rd in run_dicts:
            epoch_sets.append(set(rd.get(m, {}).keys()))

        if use_epoch_intersection:
            common = None
            for s in epoch_sets:
                common = s if common is None else common.intersection(s)
            epochs = sorted(list(common)) if common is not None else []
        else:
            union = set()
            for s in epoch_sets:
                union.update(s)
            epochs = sorted(list(union))

        if len(epochs) == 0:
            continue

        mean_time, mean_val, std_val, n_list = [], [], [], []
        for ep in epochs:
            times, vals = [], []
            for rd in run_dicts:
                if m in rd and ep in rd[m]:
                    rec = rd[m][ep]
                    t = rec.get("time", np.nan)
                    v = rec.get(value_field, np.nan)
                    if np.isfinite(t):
                        times.append(t)
                    if np.isfinite(v):
                        vals.append(v)

            n = len(vals)
            n_list.append(n)

            mean_time.append(float(np.mean(times)) if len(times) else np.nan)

            if len(vals):
                mean_val.append(float(np.mean(vals)))
                std_val.append(float(np.std(vals, ddof=1)) if len(vals) > 1 else 0.0)
            else:
                mean_val.append(np.nan)
                std_val.append(np.nan)

        agg[m] = {
            "epochs": np.array(epochs, dtype=np.int64),
            "mean_time": np.array(mean_time, dtype=np.float64),
            "mean_value": np.array(mean_val, dtype=np.float64),
            "std_value": np.array(std_val, dtype=np.float64),
            "n": np.array(n_list, dtype=np.int64),
        }

    return agg


def save_agg_tsv(agg, out_path):
    outdir = os.path.dirname(out_path)
    if outdir and (not os.path.exists(outdir)):
        os.makedirs(outdir)

    with open(out_path, "w") as f:
        writer = csv.writer(f, delimiter="\t")
        writer.writerow(["Model", "Epoch", "MeanTime(s)", "MeanTestAcc", "StdTestAcc", "N"])
        for model in sorted(agg.keys()):
            a = agg[model]
            for i in range(len(a["epochs"])):
                writer.writerow([
                    model,
                    int(a["epochs"][i]),
                    ("%.6f" % a["mean_time"][i]) if not np.isnan(a["mean_time"][i]) else "nan",
                    ("%.6f" % a["mean_acc"][i]) if not np.isnan(a["mean_acc"][i]) else "nan",
                    ("%.6f" % a["std_acc"][i]) if not np.isnan(a["std_acc"][i]) else "nan",
                    int(a["n"][i]),
                ])



def compute_final_metrics(run_dicts, models=None, last_epochs=None, last_k=5):
    """
    计算每个模型的“最终性能”：
      - 每个 run：
          * 若 last_epochs is None：自动取最后 last_k 个有 TestAcc 的 epoch 求平均
          * 否则：取 last_epochs 中存在且有 TestAcc 的 epoch 求平均
      - 跨 run：对上述平均值求 mean/std
    同时统计总训练时长：
      - 每个 run：取该模型最后一个有效 epoch 的 Time(s)
      - 跨 run：求 mean/std
    返回：
      per_run_rows: list[dict]，每行一条 (run, model) 记录
      summary_rows: list[dict]，每行一条 (model) 汇总记录
    """
    per_run_rows = []

    # 先确定模型集合
    all_models = set()
    for rd in run_dicts:
        all_models.update(rd.keys())

    if models is not None:
        use_models = [m for m in models if m in all_models]
    else:
        use_models = sorted(list(all_models))

    # per-run 计算
    for run_idx, rd in enumerate(run_dicts, start=1):
        for m in use_models:
            ep_map = rd.get(m, {})  # {epoch: (time, test_acc)}

            # ---------- Final5 TestAcc ----------
            if last_epochs is None:
                # 自动取最后 last_k 个有 TestAcc 的 epoch
                valid_acc_eps = []
                for ep, (_t, _a) in ep_map.items():
                    if np.isfinite(_a):
                        valid_acc_eps.append(ep)
                valid_acc_eps = sorted(valid_acc_eps)
                use_eps = valid_acc_eps[-last_k:]
            else:
                use_eps = []
                for ep in last_epochs:
                    if ep in ep_map:
                        _t, _a = ep_map[ep]
                        if np.isfinite(_a):
                            use_eps.append(ep)

            accs = []
            for ep in use_eps:
                _t, _a = ep_map[ep]
                if np.isfinite(_a):
                    accs.append(_a)

            final5 = float(np.mean(accs)) if len(accs) else np.nan

            # ---------- Total Time(s) ----------
            # 取该模型最后一个有有效 Time(s) 的 epoch
            total_time_s = np.nan
            valid_time_eps = []
            for ep, (_t, _a) in ep_map.items():
                if np.isfinite(_t):
                    valid_time_eps.append(ep)

            if len(valid_time_eps):
                last_ep = max(valid_time_eps)
                _t_last, _a_last = ep_map[last_ep]
                total_time_s = float(_t_last)
                #print("[DEBUG] run=%d model=%s last_ep=%d total_time_s=%.2f" % (run_idx, m, last_ep, total_time_s))

            per_run_rows.append({
                "Run": run_idx,
                "Model": m,
                "Final5_TestAcc": final5,
                "TotalTime(s)": total_time_s,
            })

    # 按模型分组
    per_run = {}
    for row in per_run_rows:
        per_run.setdefault(row["Model"], []).append(row)

    # 汇总：跨 run mean/std
    summary_rows = []
    for m in use_models:
        rows = per_run.get(m, [])

        acc_vals = [r["Final5_TestAcc"] for r in rows if np.isfinite(r["Final5_TestAcc"])]
        t_vals = [r["TotalTime(s)"] for r in rows if np.isfinite(r["TotalTime(s)"])]

        acc_mean = float(np.mean(acc_vals)) if len(acc_vals) else np.nan
        acc_std = float(np.std(acc_vals, ddof=1)) if len(acc_vals) > 1 else (0.0 if len(acc_vals) == 1 else np.nan)

        t_mean = float(np.mean(t_vals)) if len(t_vals) else np.nan
        t_std = float(np.std(t_vals, ddof=1)) if len(t_vals) > 1 else (0.0 if len(t_vals) == 1 else np.nan)

        summary_rows.append({
            "Model": m,
            "Final5_TestAcc_mean": acc_mean,
            "Final5_TestAcc_std": acc_std,
            "TotalTime(s)_mean": t_mean,
            "TotalTime(s)_std": t_std,
            "TotalTime(min)_mean": (t_mean / 60.0) if np.isfinite(t_mean) else np.nan,
            "TotalTime(min)_std": (t_std / 60.0) if np.isfinite(t_std) else np.nan,
        })

    return per_run_rows, summary_rows


def _get_ref_target_acc(ep_map, mode="final5", last_k=5):
    """给定某个模型在单个 run 的 ep_map={epoch:(time, test_acc)}，计算目标准确率。"""
    # 收集所有有效 test_acc 的 epoch
    valid_eps = []
    for ep, (_t, _a) in ep_map.items():
        if np.isfinite(_a):
            valid_eps.append(ep)
    if not valid_eps:
        return np.nan

    valid_eps = sorted(valid_eps)

    if mode == "last":
        last_ep = valid_eps[-1]
        return float(ep_map[last_ep][1])

    if mode == "best":
        best = -np.inf
        for ep in valid_eps:
            _a = ep_map[ep][1]
            if np.isfinite(_a) and _a > best:
                best = _a
        return float(best) if np.isfinite(best) else np.nan

    # 默认 final5：最后 last_k 个有效 epoch 的平均
    use_eps = valid_eps[-int(last_k):] if last_k is not None else valid_eps
    accs = [ep_map[ep][1] for ep in use_eps if np.isfinite(ep_map[ep][1])]
    return float(np.mean(accs)) if len(accs) else np.nan


def _find_first_streak_time(ep_map, target_acc, streak=1, use_end_of_streak=True, acc_eps=0.0):
    """在单个 run 内，找最早达到 target_acc 的时间点。

    返回：
      (match_epoch, match_time_s)；若未达到返回 (None, nan)
    规则：
      - 只考虑 time 与 acc 都是 finite 的 epoch
      - streak>1 时要求 epoch 连续（ep, ep+1, ..., ep+streak-1）都满足 acc >= target_acc - acc_eps
      - use_end_of_streak=True 则取窗口末尾 epoch 的时间（更保守）
    """
    if (not np.isfinite(target_acc)) or (streak is None) or (int(streak) <= 0):
        return None, np.nan

    streak = int(streak)
    # 可用 epoch 集合
    valid = set()
    for ep, (t, a) in ep_map.items():
        if np.isfinite(t) and np.isfinite(a):
            valid.add(int(ep))

    if not valid:
        return None, np.nan

    eps_sorted = sorted(valid)
    valid_set = set(eps_sorted)

    thr = float(target_acc) - float(acc_eps if acc_eps is not None else 0.0)

    for ep in eps_sorted:
        ok = True
        for k in range(streak):
            epk = ep + k
            if epk not in valid_set:
                ok = False
                break
            _t, _a = ep_map[epk]
            if (not np.isfinite(_a)) or (_a < thr):
                ok = False
                break

        if ok:
            chosen_ep = ep + streak - 1 if use_end_of_streak else ep
            chosen_t, _ = ep_map[chosen_ep]
            return int(chosen_ep), float(chosen_t)

    return None, np.nan


def compute_time_to_match_ref(
    run_dicts,
    ref_model="c",
    compare_models=None,
    ref_acc_mode="final5",
    ref_last_k=5,
    streak=3,
    use_end_of_streak=True,
    acc_eps=0.0,
):
    """统计“各模型达到 ref_model 同等 TestAcc 所需时间”。

    计算方式（每个 run）：
      1) 先用 ref_model 的 ep_map 计算目标准确率 target_acc（由 ref_acc_mode 控制）
      2) 对 compare_models 中每个模型，找最早满足 acc >= target_acc 的时间点
         - 可选连续 streak 个 epoch 都满足（抗抖动）

    返回：
      per_run_rows: 每条记录是 (run, model) 的 match 结果
      summary_rows: 按 model 汇总的 mean/std（只对“达到”的 run 统计），并给出 N_reached/N_total
    """
    per_run_rows = []

    # 确定 compare_models
    all_models = set()
    for rd in run_dicts:
        all_models.update(rd.keys())

    if compare_models is None:
        compare_models = [m for m in sorted(all_models) if m != ref_model]

    # 每个 run 先算一次 target_acc
    ref_targets = []
    for run_idx, rd in enumerate(run_dicts, start=1):
        ref_ep_map = rd.get(ref_model, {})
        target_acc = _get_ref_target_acc(ref_ep_map, mode=ref_acc_mode, last_k=ref_last_k)
        ref_targets.append(target_acc)

        for m in compare_models:
            ep_map = rd.get(m, {})
            match_ep, match_time_s = _find_first_streak_time(
                ep_map,
                target_acc,
                streak=streak,
                use_end_of_streak=use_end_of_streak,
                acc_eps=acc_eps,
            )

            per_run_rows.append({
                "Run": run_idx,
                "RefModel": ref_model,
                "RefTargetAcc": float(target_acc) if np.isfinite(target_acc) else np.nan,
                "Model": m,
                "MatchEpoch": int(match_ep) if match_ep is not None else "",
                "MatchTime(s)": float(match_time_s) if np.isfinite(match_time_s) else np.nan,
                "MatchTime(min)": (float(match_time_s) / 60.0) if np.isfinite(match_time_s) else np.nan,
                "Reached": bool(np.isfinite(match_time_s)),
            })

    # 汇总 target_acc（给参考用）
    ref_targets_valid = [x for x in ref_targets if np.isfinite(x)]
    ref_target_mean = float(np.mean(ref_targets_valid)) if len(ref_targets_valid) else np.nan
    ref_target_std = float(np.std(ref_targets_valid, ddof=1)) if len(ref_targets_valid) > 1 else (0.0 if len(ref_targets_valid) == 1 else np.nan)

    # 按模型汇总 match time
    by_model = {}
    for row in per_run_rows:
        by_model.setdefault(row["Model"], []).append(row)

    summary_rows = []
    n_total = len(run_dicts)

    for m in compare_models:
        rows = by_model.get(m, [])
        times_s = [r["MatchTime(s)"] for r in rows if np.isfinite(r["MatchTime(s)"])]
        n_reached = len(times_s)

        mean_s = float(np.mean(times_s)) if n_reached else np.nan
        std_s = float(np.std(times_s, ddof=1)) if n_reached > 1 else (0.0 if n_reached == 1 else np.nan)

        summary_rows.append({
            "Model": m,
            "TimeToMatchRef(s)_mean": mean_s,
            "TimeToMatchRef(s)_std": std_s,
            "TimeToMatchRef(min)_mean": (mean_s / 60.0) if np.isfinite(mean_s) else np.nan,
            "TimeToMatchRef(min)_std": (std_s / 60.0) if np.isfinite(std_s) else np.nan,
            "N_reached": int(n_reached),
            "N_total": int(n_total),
            "RefModel": ref_model,
            "RefTargetAcc_mean": ref_target_mean,
            "RefTargetAcc_std": ref_target_std,
            "RefAccMode": str(ref_acc_mode),
            "Streak": int(streak),
            "UseEndOfStreak": bool(use_end_of_streak),
            "AccEps": float(acc_eps if acc_eps is not None else 0.0),
        })

    return per_run_rows, summary_rows


def save_match_time_csv(summary_rows, out_csv_path):
    """保存“达到 Big Net 同等 TestAcc 所需时间”的汇总 CSV。"""
    outdir = os.path.dirname(out_csv_path)
    if outdir and (not os.path.exists(outdir)):
        os.makedirs(outdir)

    with open(out_csv_path, "w") as f:
        writer = csv.writer(f)
        writer.writerow([
            "Model",
            "TimeToMatchRef(min)_mean",
            "TimeToMatchRef(min)_std",
            "TimeToMatchRef(s)_mean",
            "TimeToMatchRef(s)_std",
            "N_reached",
            "N_total",
            "RefModel",
            "RefTargetAcc_mean",
            "RefTargetAcc_std",
            "RefAccMode",
            "Streak",
            "UseEndOfStreak",
            "AccEps",
        ])

        # 固定输出顺序：a,b,d,e,f（如果存在）
        models = [r["Model"] for r in summary_rows]
        prefer = ["a", "b", "d", "e", "f"]
        order = [m for m in prefer if m in models] + [m for m in sorted(models) if m not in prefer]

        row_map = {r["Model"]: r for r in summary_rows}
        for m in order:
            r = row_map[m]
            writer.writerow([
                r["Model"],
                ("%.4f" % r["TimeToMatchRef(min)_mean"]) if np.isfinite(r["TimeToMatchRef(min)_mean"]) else "nan",
                ("%.4f" % r["TimeToMatchRef(min)_std"]) if np.isfinite(r["TimeToMatchRef(min)_std"]) else "nan",
                ("%.6f" % r["TimeToMatchRef(s)_mean"]) if np.isfinite(r["TimeToMatchRef(s)_mean"]) else "nan",
                ("%.6f" % r["TimeToMatchRef(s)_std"]) if np.isfinite(r["TimeToMatchRef(s)_std"]) else "nan",
                int(r["N_reached"]),
                int(r["N_total"]),
                r["RefModel"],
                ("%.6f" % r["RefTargetAcc_mean"]) if np.isfinite(r["RefTargetAcc_mean"]) else "nan",
                ("%.6f" % r["RefTargetAcc_std"]) if np.isfinite(r["RefTargetAcc_std"]) else "nan",
                r["RefAccMode"],
                int(r["Streak"]),
                int(1 if r["UseEndOfStreak"] else 0),
                ("%.6f" % r["AccEps"]) if np.isfinite(r["AccEps"]) else "nan",
            ])


def save_match_time_excel(summary_rows, out_xlsx_path):
    """生成一个 Excel（类似 final_testacc_summary_formatted.xlsx 的排版），展示达到目标 TestAcc 的时间。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise ImportError('缺少 openpyxl：请 pip install openpyxl') from e

    outdir = os.path.dirname(out_xlsx_path)
    if outdir and (not os.path.exists(outdir)):
        os.makedirs(outdir)

    display_name_map = {
        "c": "Big Net",
        "a": "Mode a",
        "b": "Mode b",
        "d": "Mode c",
        "e": "Mode d",
        "f": "Gradmax",
    }

    models = [r["Model"] for r in summary_rows]
    prefer = ["a", "b", "d", "e", "f"]
    order = [m for m in prefer if m in models] + [m for m in sorted(models) if m not in prefer]

    row_map = {r["Model"]: r for r in summary_rows}

    # 从任意一行取 ref 相关信息（这些对所有模型相同）
    ref_model = summary_rows[0]["RefModel"] if len(summary_rows) else "c"
    ref_mean = summary_rows[0]["RefTargetAcc_mean"] if len(summary_rows) else np.nan
    ref_std = summary_rows[0]["RefTargetAcc_std"] if len(summary_rows) else np.nan
    ref_mode = summary_rows[0]["RefAccMode"] if len(summary_rows) else "final5"
    streak = summary_rows[0]["Streak"] if len(summary_rows) else 1
    use_end = summary_rows[0]["UseEndOfStreak"] if len(summary_rows) else True

    def _fmt(mean_v, std_v, mean_dec=2, std_dec=3):
        if not np.isfinite(mean_v):
            return ''
        if not np.isfinite(std_v):
            std_v = 0.0
        return ('%.' + str(mean_dec) + 'f ± %.' + str(std_dec) + 'f') % (mean_v, std_v)

    header = ["mean ± std over 3 runs"] + [display_name_map.get(m, "Model %s" % m) for m in order]

    # 行1：目标准确率（只在 A 列解释，B 列给出值，其余留空）
    target_label = "Target TestAcc (%s, %s)" % (display_name_map.get(ref_model, ref_model), ref_mode)
    target_row = [target_label] + ["%.6f ± %.6f" % (ref_mean, ref_std) if np.isfinite(ref_mean) else ""]

    # 对齐到列数
    while len(target_row) < len(header):
        target_row.append("")

    time_label = "Time to reach target (mins, streak=%d, %s)" % (int(streak), "end" if use_end else "start")
    time_row = [time_label] + [
        _fmt(row_map[m]["TimeToMatchRef(min)_mean"], row_map[m]["TimeToMatchRef(min)_std"], 2, 3)
        for m in order
    ]

    reach_label = "Reached runs (N/Total)"
    reach_row = [reach_label] + [
        ("%d/%d" % (int(row_map[m]["N_reached"]), int(row_map[m]["N_total"]))) for m in order
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    # 写入
    for j, v in enumerate(header, start=1):
        ws.cell(row=1, column=j, value=v)
    for j, v in enumerate(target_row, start=1):
        ws.cell(row=2, column=j, value=v)
    for j, v in enumerate(time_row, start=1):
        ws.cell(row=3, column=j, value=v)
    for j, v in enumerate(reach_row, start=1):
        ws.cell(row=4, column=j, value=v)

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # header 样式
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = center
        cell.border = border

    # 内容样式
    for r in [2, 3, 4]:
        ws.cell(row=r, column=1).alignment = left
        ws.cell(row=r, column=1).border = border
        for c in range(2, len(header) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = center
            cell.border = border

    ws.column_dimensions["A"].width = 34
    for idx in range(2, len(header) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 16

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 20

    wb.save(out_xlsx_path)


def save_final_metrics_csv(summary_rows, out_csv_path):
    """保存最终性能汇总 CSV（包含 mean/std + 总训练时长）。"""
    outdir = os.path.dirname(out_csv_path)
    if outdir and (not os.path.exists(outdir)):
        os.makedirs(outdir)

    with open(out_csv_path, "w") as f:
        writer = csv.writer(f)
        writer.writerow([
            "Model",
            "Final5_TestAcc_mean",
            "Final5_TestAcc_std",
            "TotalTime(min)_mean",
            "TotalTime(min)_std",
            "TotalTime(s)_mean",
            "TotalTime(s)_std",
        ])
        models = [r["Model"] for r in summary_rows]
        order = []
        if "c" in models:
            order.append("c")
        order += [m for m in sorted(models) if m != "c"]
        row_map = {r["Model"]: r for r in summary_rows}
        for m in order:
            row = row_map[m]
            writer.writerow([
                row["Model"],
                ("%.6f" % row["Final5_TestAcc_mean"]) if np.isfinite(row["Final5_TestAcc_mean"]) else "nan",
                ("%.6f" % row["Final5_TestAcc_std"]) if np.isfinite(row["Final5_TestAcc_std"]) else "nan",
                ("%.4f" % row["TotalTime(min)_mean"]) if np.isfinite(row["TotalTime(min)_mean"]) else "nan",
                ("%.4f" % row["TotalTime(min)_std"]) if np.isfinite(row["TotalTime(min)_std"]) else "nan",
                ("%.6f" % row["TotalTime(s)_mean"]) if np.isfinite(row["TotalTime(s)_mean"]) else "nan",
                ("%.6f" % row["TotalTime(s)_std"]) if np.isfinite(row["TotalTime(s)_std"]) else "nan",
            ])

def save_final_metrics_excel(summary_rows, out_xlsx_path):
    """
    额外生成一个 Excel（排版类似你截图）：
      - 第一行：mean ± std over 3 runs + 各模型显示名
      - 第一列：TestAcc / Time (mins)
      - 单元格内容：mean ± std（格式化为字符串）
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise ImportError('缺少 openpyxl：请 pip install openpyxl') from e

    outdir = os.path.dirname(out_xlsx_path)
    if outdir and (not os.path.exists(outdir)):
        os.makedirs(outdir)

    # 与其他图一致的显示名（按你截图：Mode a/b/c/d）
    display_name_map = {
        "c": "Big Net",
        "a": "Mode a",
        "b": "Mode b",
        "d": "Mode c",
        "e": "Mode d",
        "f": "Gradmax",
    }

    models = [r['Model'] for r in summary_rows]
    order = []
    if 'c' in models:
        order.append('c')
    order += [m for m in ['a','b','d','e','f'] if (m in models and m != 'c')]
    order += [m for m in sorted(models) if m not in order]

    # 取值映射
    row_map = {r['Model']: r for r in summary_rows}

    def _fmt(mean_v, std_v, mean_dec=3, std_dec=4):
        if not np.isfinite(mean_v):
            return ''
        if not np.isfinite(std_v):
            std_v = 0.0
        return ('%.' + str(mean_dec) + 'f ± %.' + str(std_dec) + 'f') % (mean_v, std_v)

    header = ['mean ± std over 3 runs'] + [display_name_map.get(m, 'Model %s' % m) for m in order]
    acc_row = ['TestAcc'] + [_fmt(row_map[m]['Final5_TestAcc_mean'], row_map[m]['Final5_TestAcc_std'], 3, 4) for m in order]
    time_row = ['Time (mins)'] + [_fmt(row_map[m]['TotalTime(min)_mean'], row_map[m]['TotalTime(min)_std'], 2, 3) for m in order]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Summary'

    for j, v in enumerate(header, start=1):
        ws.cell(row=1, column=j, value=v)
    for j, v in enumerate(acc_row, start=1):
        ws.cell(row=2, column=j, value=v)
    for j, v in enumerate(time_row, start=1):
        ws.cell(row=3, column=j, value=v)

    bold = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')
    thin = Side(style='thin', color='D0D0D0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font = bold
        cell.alignment = center
        cell.border = border

    for r in [2, 3]:
        ws.cell(row=r, column=1).alignment = left
        ws.cell(row=r, column=1).border = border
        for c in range(2, len(header) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = center
            cell.border = border

    # 列宽
    ws.column_dimensions['A'].width = 18
    for idx in range(2, len(header) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 14
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20

    wb.save(out_xlsx_path)
    
def plot_final_acc_bar(summary_rows, out_png, title="Final TestAcc (5 last Epochs mean over runs)"):
    models = [r["Model"] for r in summary_rows]
    order = []
    if "c" in models:
        order.append("c")
    order += [m for m in sorted(models) if m != "c"]

    mean_map = {r["Model"]: r["Final5_TestAcc_mean"] for r in summary_rows}
    std_map = {r["Model"]: r["Final5_TestAcc_std"] for r in summary_rows}

    xs = np.arange(len(order), dtype=np.float64)
    ys = np.array([mean_map.get(m, np.nan) for m in order], dtype=np.float64)
    es = np.array([std_map.get(m, np.nan) for m in order], dtype=np.float64)

    mask = np.isfinite(ys)
    xs2 = xs[mask]
    order2 = [m for (m, ok) in zip(order, mask) if ok]
    ys2 = ys[mask]
    es2 = es[mask]
    es2 = np.where(np.isfinite(es2), es2, 0.0)

    color_map = {
        "c": "green",
        "a": "blue",
        "b": "orange",
        "d": "brown",
        "e": "deeppink",
        "f": "purple",
    }

    bar_colors = [color_map.get(m, None) for m in order2]

    plt.figure()
    plt.bar(xs2, ys2, color=bar_colors)
    plt.errorbar(xs2, ys2, yerr=es2, fmt="none",
             capsize=10, capthick=2.4, elinewidth=2.4,
             ecolor="black", zorder=3)

    display_name_map = {
        "c": "Big Net",
        "a": "model a",
        "b": "model b",
        "d": "model c",
        "e": "model d",
        "f": "Gradmax",
    }

    plt.xticks(xs2, [display_name_map.get(m, "Model %s" % m) for m in order2])
    plt.ylim(0.85, 0.92)
    plt.xlabel("Model")
    plt.ylabel("TestAcc (mean over runs)")
    plt.title(title)
    plt.grid(True, axis="y")
    plt.tight_layout()

    outdir = os.path.dirname(out_png)
    if outdir and (not os.path.exists(outdir)):
        os.makedirs(outdir)

    plt.savefig(out_png)
    plt.close()


def plot_mean_std(agg, out_png):
    color_map = {
        "c": "green",
        "a": "blue",
        "b": "orange",
        "d": "brown",
        "e": "deeppink",
        "f": "purple",
    }

    plt.figure()
    
    plot_order = []
    if "c" in agg:
        plot_order.append("c")
    plot_order += [m for m in sorted(agg.keys()) if m != "c"]

    for model in plot_order:
        a = agg[model]
        x = a["mean_time"]
        y = a["mean_acc"]
        s = a["std_acc"]

        mask = np.isfinite(x) & np.isfinite(y) & np.isfinite(s)
        x, y, s = x[mask], y[mask], s[mask]
        if x.size == 0:
            continue

        color = color_map.get(model, None)
        display_name = {
            "c": "Big Net",
            "a": "model a",
            "b": "model b",
            "d": "model c",
            "e": "model d",
            "f": "Gradmax",
        }.get(model, "Model %s" % model)
        label = display_name

        plt.plot(x, y, label=label, color=color)
        plt.fill_between(x, y - s, y + s, color=color, alpha=0.2)

    plt.xlabel("Time (s)")
    plt.ylabel("Test Accuracy")
    #plt.ylim(0.980, 0.987) 
    plt.title(FIG_TITLE)
    plt.legend()
    plt.grid(True)
    plt.tight_layout()

    outdir = os.path.dirname(out_png)
    if outdir and (not os.path.exists(outdir)):
        os.makedirs(outdir)

    plt.savefig(out_png)
    plt.close()


def plot_metric_mean_std(agg, out_png, ylabel, title):
    """画：某个指标 vs Time 的均值曲线 + STD 阴影（风格与 TestAcc 图一致）。"""
    color_map = {
        "c": "green",
        "a": "blue",
        "b": "orange",
        "d": "brown",
        "e": "deeppink",
        "f": "purple",
    }

    plt.figure()

    plot_order = []
    if "c" in agg:
        plot_order.append("c")
    plot_order += [m for m in sorted(agg.keys()) if m != "c"]

    plotted = 0
    for model in plot_order:
        a = agg[model]
        x = a["mean_time"]
        y = a["mean_value"]
        s = a["std_value"]

        mask = np.isfinite(x) & np.isfinite(y) & np.isfinite(s)
        x, y, s = x[mask], y[mask], s[mask]
        if x.size == 0:
            continue

        color = color_map.get(model, None)
        display_name = {
            "c": "Big Net",
            "a": "model a",
            "b": "model b",
            "d": "model c",
            "e": "model d",
            "f": "Gradmax",
        }.get(model, "Model %s" % model)

        plt.plot(x, y, label=display_name, color=color)
        plt.fill_between(x, y - s, y + s, color=color, alpha=0.2)
        plotted += 1

    if plotted == 0:
        raise ValueError("没有可画的数据点：%s" % title)

    plt.xlabel("Time (s)")
    plt.ylabel(ylabel)
    plt.title(title)
    plt.legend()
    plt.grid(True)
    plt.tight_layout()

    outdir = os.path.dirname(out_png)
    if outdir and (not os.path.exists(outdir)):
        os.makedirs(outdir)

    plt.savefig(out_png)
    plt.close()


def plot_mean_std_epoch(agg, out_png):
    """画：Test Accuracy vs Epoch 的均值曲线 + STD 阴影。"""
    color_map = {
        "c": "green",
        "a": "blue",
        "b": "orange",
        "d": "brown",
        "e": "deeppink",
        "f": "purple",
    }

    plt.figure()

    # 让 Big Net (model 'c') 在 legend 最上面：先画 'c'，再画其他
    plot_order = []
    if "c" in agg:
        plot_order.append("c")
    plot_order += [m for m in sorted(agg.keys()) if m != "c"]

    for model in plot_order:
        a = agg[model]
        x = a["epochs"].astype(np.float64)
        y = a["mean_acc"]
        s = a["std_acc"]

        mask = np.isfinite(x) & np.isfinite(y) & np.isfinite(s)
        x, y, s = x[mask], y[mask], s[mask]
        if x.size == 0:
            continue

        color = color_map.get(model, None)
        display_name = {
            "c": "Big Net",
            "a": "model a",
            "b": "model b",
            "d": "model c",
            "e": "model d",
            "f": "Gradmax",
        }.get(model, "Model %s" % model)
        label = display_name

        plt.plot(x, y, label=label, color=color)
        plt.fill_between(x, y - s, y + s, color=color, alpha=0.2)

    plt.xlabel("Epoch")
    plt.ylabel("Test Accuracy")
    plt.title("Test Accuracy vs Epoch (mean ± std over 3 runs)")
    plt.legend()
    plt.grid(True)
    plt.tight_layout()

    outdir = os.path.dirname(out_png)
    if outdir and (not os.path.exists(outdir)):
        os.makedirs(outdir)

    plt.savefig(out_png)
    plt.close()


def main():
    # 检查输入文件
    for p in INPUT_FILES:
        if not os.path.isfile(p):
            raise IOError("找不到输入文件: %s" % p)

    run_dicts = [read_results_tsv(p) for p in INPUT_FILES]
    run_dicts_all = [read_results_tsv_allmetrics(p) for p in INPUT_FILES]

    agg = aggregate_runs(
        run_dicts,
        models=MODELS_TO_PLOT,
        use_epoch_intersection=USE_EPOCH_INTERSECTION,
    )
    if len(agg) == 0:
        raise ValueError("没有可画的数据：请检查 MODELS_TO_PLOT 或 results 文件内容。")

    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
        
    out_tsv = os.path.join(OUTPUT_DIR, "test_accuracy_vs_time_mean.tsv")

    """
    out_png = os.path.join(OUTPUT_DIR, "test_accuracy_vs_time_mean.png")
    out_png_epoch = os.path.join(OUTPUT_DIR, "test_accuracy_vs_epoch_mean.png")
    
    
    # ===== 新增：三张“随时间变化”的曲线图（mean ± std over 3 runs） =====
    out_png_train_loss = os.path.join(OUTPUT_DIR, "train_loss_vs_time_mean.png")
    out_png_train_acc = os.path.join(OUTPUT_DIR, "train_accuracy_vs_time_mean.png")
    out_png_test_loss = os.path.join(OUTPUT_DIR, "test_loss_vs_time_mean.png")

    agg_train_loss = aggregate_runs_metric(
        run_dicts_all,
        value_field="train_loss",
        models=MODELS_TO_PLOT,
        use_epoch_intersection=USE_EPOCH_INTERSECTION,
    )
    agg_train_acc = aggregate_runs_metric(
        run_dicts_all,
        value_field="train_acc",
        models=MODELS_TO_PLOT,
        use_epoch_intersection=USE_EPOCH_INTERSECTION,
    )
    agg_test_loss = aggregate_runs_metric(
        run_dicts_all,
        value_field="test_loss",
        models=MODELS_TO_PLOT,
        use_epoch_intersection=USE_EPOCH_INTERSECTION,
    )

    # ===== 新增：最终性能(最后5个epoch)汇总 + CSV + 柱状图 =====
    final_per_run_rows, final_summary_rows = compute_final_metrics(
        run_dicts,
        models=MODELS_TO_PLOT,
        last_epochs=None,   # 自动取最后5个有效epoch
        last_k=5,
    )
    
    out_final_csv = os.path.join(OUTPUT_DIR, "final_testacc_summary.csv")
    out_final_bar = os.path.join(OUTPUT_DIR, "final_testacc_bar_mean_std.png")
    save_final_metrics_csv(final_summary_rows, out_final_csv)
    out_final_xlsx = os.path.join(OUTPUT_DIR, "final_testacc_summary_formatted.xlsx")
    save_final_metrics_excel(final_summary_rows, out_final_xlsx)
    plot_final_acc_bar(final_summary_rows, out_final_bar)
    """

    # ===== 新增：统计“达到 Big Net(c) 同等 TestAcc 所需时间” + CSV + Excel =====
    match_per_run_rows, match_summary_rows = compute_time_to_match_ref(
        run_dicts,
        ref_model=MATCH_REF_MODEL,
        compare_models=MATCH_COMPARE_MODELS,
        ref_acc_mode=MATCH_REF_ACC_MODE,
        ref_last_k=MATCH_REF_LAST_K,
        streak=MATCH_STREAK,
        use_end_of_streak=MATCH_USE_END_OF_STREAK,
        acc_eps=MATCH_ACC_EPS,
    )

    out_match_csv = os.path.join(OUTPUT_DIR, "time_to_match_c_summary.csv")
    out_match_xlsx = os.path.join(OUTPUT_DIR, "time_to_match_c_summary_formatted.xlsx")
    save_match_time_csv(match_summary_rows, out_match_csv)
    save_match_time_excel(match_summary_rows, out_match_xlsx)

    """
    save_agg_tsv(agg, out_tsv)
    plot_mean_std(agg, out_png)
    plot_mean_std_epoch(agg, out_png_epoch)

    # 新增三张图
    plot_metric_mean_std(
        agg_train_loss,
        out_png_train_loss,
        ylabel="Train Loss",
        title="Train Loss vs Time (mean ± std over 3 runs)",
    )
    plot_metric_mean_std(
        agg_train_acc,
        out_png_train_acc,
        ylabel="Train Accuracy",
        title="Train Accuracy vs Time (mean ± std over 3 runs)",
    )
    plot_metric_mean_std(
        agg_test_loss,
        out_png_test_loss,
        ylabel="Test Loss",
        title="Test Loss vs Time (mean ± std over 3 runs)",
    )

    print("已保存：")
    print("  图1(Time):  %s" % out_png)
    print("  图2(Epoch): %s" % out_png_epoch)
    print("  图3(TrainLoss-Time): %s" % out_png_train_loss)
    print("  图4(TrainAcc-Time):  %s" % out_png_train_acc)
    print("  图5(TestLoss-Time):  %s" % out_png_test_loss)
    print("  表:         %s" % out_tsv)
    print("  最终性能CSV: %s" % out_final_csv)
    print("  柱状图:      %s" % out_final_bar)
    """
    print("  达标时间CSV: %s" % out_match_csv)
    print("  达标时间XLSX:%s" % out_match_xlsx)


if __name__ == "__main__":
    main()