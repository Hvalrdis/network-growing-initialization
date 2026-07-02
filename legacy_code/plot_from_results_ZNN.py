# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import csv
import os
from typing import Dict, List, Optional, Sequence, Tuple

import numpy as np

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt


# =========================
# 默认配置（可直接运行）
# =========================
DEFAULT_INPUT_FILES = [
    "/home/ywang/Article/vgg_gradmax/outputs_compare/vgg_four_modes/20260413_182237_run02_seed2124906467_lr0.05/results_run_2.txt",
    "/home/ywang/Article/vgg_gradmax/outputs_compare/vgg_four_modes/20260413_201324_run03_seed963822966_lr0.05/results_run_3.txt",
    "/home/ywang/Article/vgg_gradmax/outputs_compare/vgg_four_modes/20260414_121920_run01_seed5_lr0.05/results_run_1.txt"
]

DEFAULT_MODELS = ["mode_NNN", "mode_NZZ", "mode_ZNN", "mode_ZNZ"]

DISPLAY_NAME_MAP = {
    "mode_NNN": "mode_NNN",
    "mode_NZZ": "mode_NZZ",
    "mode_ZNN": "mode_ZNN",
    "mode_ZNZ": "mode_ZNZ",
}

COLOR_MAP = {
    "mode_NNN": "tab:blue",
    "mode_NZZ": "tab:orange",
    "mode_ZNN": "tab:green",
    "mode_ZNZ": "tab:red",
}

USE_EPOCH_INTERSECTION = True
FINAL_LAST_K = 5

# =========================
# 可直接在脚本里改的 ylim 配置
# 写成 None 表示不限制；写成 (ymin, ymax) 表示固定纵轴范围
# =========================
YLIM_TEST_ACC_TIME = None
YLIM_TEST_ACC_EPOCH = None
YLIM_TRAIN_LOSS_TIME = None
YLIM_TRAIN_ACC_TIME = None
YLIM_TEST_LOSS_TIME = None
YLIM_PARAM_COUNT_EPOCH = None
YLIM_FINAL_ACC_BAR = (0.85, 0.88)


# =========================
# 工具函数
# =========================
def _safe_float(x, default=np.nan):
    try:
        return float(x)
    except Exception:
        return default


def _safe_int(x, default=None):
    try:
        return int(x)
    except Exception:
        return default


def _ensure_dir(path: str) -> None:
    if path and (not os.path.exists(path)):
        os.makedirs(path)


def _default_output_dir(input_files: List[str]) -> str:
    common = os.path.commonpath(input_files)
    if os.path.isfile(common):
        common = os.path.dirname(common)
    return os.path.join(common, "summary_four_modes")


def _parse_ylim_arg(values: Optional[Sequence[float]]) -> Optional[Tuple[float, float]]:
    if values is None:
        return None
    if len(values) != 2:
        raise ValueError(f"ylim 参数必须正好给两个值，收到: {values}")
    y0, y1 = float(values[0]), float(values[1])
    if y0 >= y1:
        raise ValueError(f"ylim 需要满足 ymin < ymax，收到: {(y0, y1)}")
    return (y0, y1)


def _resolve_ylim(cli_ylim: Optional[Tuple[float, float]], default_ylim: Optional[Tuple[float, float]]) -> Optional[Tuple[float, float]]:
    return cli_ylim if cli_ylim is not None else default_ylim


def _apply_ylim(ylim: Optional[Tuple[float, float]]) -> None:
    if ylim is None:
        return
    plt.ylim(float(ylim[0]), float(ylim[1]))


# =========================
# 读取结果文件
# =========================
def read_results_tsv_allmetrics(path: str) -> Dict[str, Dict[int, dict]]:
    """读取单个 results_run_*.txt。

    返回：
        data[model][epoch] = {
            'time': float,
            'train_loss': float,
            'train_acc': float,
            'test_loss': float,
            'test_acc': float,
            'param_count': float,
        }
    """
    data: Dict[str, Dict[int, dict]] = {}
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter="\t")
        required = ["Epoch", "Model", "Time(s)"]
        for key in required:
            if key not in reader.fieldnames:
                raise ValueError(f"文件缺少列 {key}，实际列为 {reader.fieldnames} -> {path}")

        for row in reader:
            epoch = _safe_int(row.get("Epoch"))
            model = row.get("Model")
            if epoch is None or model is None:
                continue

            rec = {
                "time": _safe_float(row.get("Time(s)")),
                "train_loss": _safe_float(row.get("TrainLoss")),
                "train_acc": _safe_float(row.get("TrainAcc")),
                "test_loss": _safe_float(row.get("TestLoss")),
                "test_acc": _safe_float(row.get("TestAcc")),
                "param_count": _safe_float(row.get("ParamCount")),
            }
            data.setdefault(model, {})[epoch] = rec
    return data


# =========================
# 聚合多个 run
# =========================
def aggregate_runs_metric(
    run_dicts: List[Dict[str, Dict[int, dict]]],
    value_field: str,
    models: Optional[List[str]] = None,
    use_epoch_intersection: bool = True,
) -> Dict[str, dict]:
    all_models = set()
    for rd in run_dicts:
        all_models.update(rd.keys())

    if models is None:
        use_models = sorted(all_models)
    else:
        use_models = [m for m in models if m in all_models]

    agg: Dict[str, dict] = {}
    for model in use_models:
        epoch_sets = [set(rd.get(model, {}).keys()) for rd in run_dicts]
        if use_epoch_intersection:
            common = None
            for s in epoch_sets:
                common = s if common is None else common.intersection(s)
            epochs = sorted(common) if common is not None else []
        else:
            union = set()
            for s in epoch_sets:
                union.update(s)
            epochs = sorted(union)

        if not epochs:
            continue

        mean_time, std_time = [], []
        mean_value, std_value = [], []
        n_list = []

        for ep in epochs:
            times, vals = [], []
            for rd in run_dicts:
                if model in rd and ep in rd[model]:
                    rec = rd[model][ep]
                    t = rec.get("time", np.nan)
                    v = rec.get(value_field, np.nan)
                    if np.isfinite(t):
                        times.append(t)
                    if np.isfinite(v):
                        vals.append(v)

            n = len(vals)
            n_list.append(n)
            if times:
                mean_time.append(float(np.mean(times)))
                std_time.append(float(np.std(times, ddof=1)) if len(times) > 1 else 0.0)
            else:
                mean_time.append(np.nan)
                std_time.append(np.nan)

            if vals:
                mean_value.append(float(np.mean(vals)))
                std_value.append(float(np.std(vals, ddof=1)) if len(vals) > 1 else 0.0)
            else:
                mean_value.append(np.nan)
                std_value.append(np.nan)

        agg[model] = {
            "epochs": np.array(epochs, dtype=np.int64),
            "mean_time": np.array(mean_time, dtype=np.float64),
            "std_time": np.array(std_time, dtype=np.float64),
            "mean_value": np.array(mean_value, dtype=np.float64),
            "std_value": np.array(std_value, dtype=np.float64),
            "n": np.array(n_list, dtype=np.int64),
        }
    return agg


def save_agg_tsv(agg: Dict[str, dict], out_path: str, value_name: str) -> None:
    _ensure_dir(os.path.dirname(out_path))
    with open(out_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter="\t")
        writer.writerow(["Model", "Epoch", "MeanTime(s)", f"Mean{value_name}", f"Std{value_name}", "N"])
        for model in DEFAULT_MODELS:
            if model not in agg:
                continue
            a = agg[model]
            for i in range(len(a["epochs"])):
                writer.writerow([
                    model,
                    int(a["epochs"][i]),
                    "%.6f" % a["mean_time"][i] if np.isfinite(a["mean_time"][i]) else "nan",
                    "%.6f" % a["mean_value"][i] if np.isfinite(a["mean_value"][i]) else "nan",
                    "%.6f" % a["std_value"][i] if np.isfinite(a["std_value"][i]) else "nan",
                    int(a["n"][i]),
                ])


# =========================
# 最终指标汇总（最后5个 epoch 平均）
# =========================
def compute_final_metrics(
    run_dicts: List[Dict[str, Dict[int, dict]]],
    models: Optional[List[str]] = None,
    last_k: int = 5,
) -> tuple[list[dict], list[dict]]:
    all_models = set()
    for rd in run_dicts:
        all_models.update(rd.keys())

    if models is None:
        use_models = sorted(all_models)
    else:
        use_models = [m for m in models if m in all_models]

    per_run_rows = []
    for run_idx, rd in enumerate(run_dicts, start=1):
        for model in use_models:
            ep_map = rd.get(model, {})

            valid_acc_eps = sorted(ep for ep, rec in ep_map.items() if np.isfinite(rec.get("test_acc", np.nan)))
            use_eps = valid_acc_eps[-int(last_k):]
            accs = [ep_map[ep]["test_acc"] for ep in use_eps if np.isfinite(ep_map[ep]["test_acc"])]
            final_acc = float(np.mean(accs)) if accs else np.nan

            valid_time_eps = sorted(ep for ep, rec in ep_map.items() if np.isfinite(rec.get("time", np.nan)))
            total_time_s = float(ep_map[valid_time_eps[-1]]["time"]) if valid_time_eps else np.nan

            valid_param_eps = sorted(ep for ep, rec in ep_map.items() if np.isfinite(rec.get("param_count", np.nan)))
            final_param = float(ep_map[valid_param_eps[-1]]["param_count"]) if valid_param_eps else np.nan

            per_run_rows.append({
                "Run": run_idx,
                "Model": model,
                "Final5_TestAcc": final_acc,
                "TotalTime(s)": total_time_s,
                "FinalParamCount": final_param,
            })

    summary_rows = []
    for model in use_models:
        rows = [r for r in per_run_rows if r["Model"] == model]
        acc_vals = [r["Final5_TestAcc"] for r in rows if np.isfinite(r["Final5_TestAcc"])]
        time_vals = [r["TotalTime(s)"] for r in rows if np.isfinite(r["TotalTime(s)"])]
        param_vals = [r["FinalParamCount"] for r in rows if np.isfinite(r["FinalParamCount"])]

        acc_mean = float(np.mean(acc_vals)) if acc_vals else np.nan
        acc_std = float(np.std(acc_vals, ddof=1)) if len(acc_vals) > 1 else (0.0 if len(acc_vals) == 1 else np.nan)
        time_mean = float(np.mean(time_vals)) if time_vals else np.nan
        time_std = float(np.std(time_vals, ddof=1)) if len(time_vals) > 1 else (0.0 if len(time_vals) == 1 else np.nan)
        param_mean = float(np.mean(param_vals)) if param_vals else np.nan
        param_std = float(np.std(param_vals, ddof=1)) if len(param_vals) > 1 else (0.0 if len(param_vals) == 1 else np.nan)

        summary_rows.append({
            "Model": model,
            "Final5_TestAcc_mean": acc_mean,
            "Final5_TestAcc_std": acc_std,
            "TotalTime(s)_mean": time_mean,
            "TotalTime(s)_std": time_std,
            "TotalTime(min)_mean": (time_mean / 60.0) if np.isfinite(time_mean) else np.nan,
            "TotalTime(min)_std": (time_std / 60.0) if np.isfinite(time_std) else np.nan,
            "FinalParamCount_mean": param_mean,
            "FinalParamCount_std": param_std,
        })

    return per_run_rows, summary_rows


def save_final_metrics_csv(per_run_rows: List[dict], summary_rows: List[dict], out_dir: str) -> None:
    _ensure_dir(out_dir)

    per_run_csv = os.path.join(out_dir, "final_metrics_per_run.csv")
    with open(per_run_csv, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Run", "Model", "Final5_TestAcc", "TotalTime(s)", "FinalParamCount"])
        for row in per_run_rows:
            writer.writerow([
                row["Run"],
                row["Model"],
                "%.6f" % row["Final5_TestAcc"] if np.isfinite(row["Final5_TestAcc"]) else "nan",
                "%.6f" % row["TotalTime(s)"] if np.isfinite(row["TotalTime(s)"]) else "nan",
                "%.0f" % row["FinalParamCount"] if np.isfinite(row["FinalParamCount"]) else "nan",
            ])

    summary_csv = os.path.join(out_dir, "final_metrics_summary.csv")
    with open(summary_csv, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([
            "Model",
            "Final5_TestAcc_mean", "Final5_TestAcc_std",
            "TotalTime(min)_mean", "TotalTime(min)_std",
            "TotalTime(s)_mean", "TotalTime(s)_std",
            "FinalParamCount_mean", "FinalParamCount_std",
        ])
        row_map = {r["Model"]: r for r in summary_rows}
        for model in DEFAULT_MODELS:
            if model not in row_map:
                continue
            r = row_map[model]
            writer.writerow([
                model,
                "%.6f" % r["Final5_TestAcc_mean"] if np.isfinite(r["Final5_TestAcc_mean"]) else "nan",
                "%.6f" % r["Final5_TestAcc_std"] if np.isfinite(r["Final5_TestAcc_std"]) else "nan",
                "%.4f" % r["TotalTime(min)_mean"] if np.isfinite(r["TotalTime(min)_mean"]) else "nan",
                "%.4f" % r["TotalTime(min)_std"] if np.isfinite(r["TotalTime(min)_std"]) else "nan",
                "%.6f" % r["TotalTime(s)_mean"] if np.isfinite(r["TotalTime(s)_mean"]) else "nan",
                "%.6f" % r["TotalTime(s)_std"] if np.isfinite(r["TotalTime(s)_std"]) else "nan",
                "%.0f" % r["FinalParamCount_mean"] if np.isfinite(r["FinalParamCount_mean"]) else "nan",
                "%.0f" % r["FinalParamCount_std"] if np.isfinite(r["FinalParamCount_std"]) else "nan",
            ])


def save_final_metrics_excel(summary_rows: List[dict], out_xlsx_path: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception as e:
        print(f"[WARN] 跳过 Excel 导出（需要 openpyxl）：{e}")
        return

    _ensure_dir(os.path.dirname(out_xlsx_path))

    def _fmt(mean_v, std_v, mean_dec=3, std_dec=4):
        if not np.isfinite(mean_v):
            return ""
        if not np.isfinite(std_v):
            std_v = 0.0
        return ("%." + str(mean_dec) + "f ± %." + str(std_dec) + "f") % (mean_v, std_v)

    row_map = {r["Model"]: r for r in summary_rows}
    header = ["mean ± std over 3 runs"] + [DISPLAY_NAME_MAP.get(m, m) for m in DEFAULT_MODELS if m in row_map]
    models = [m for m in DEFAULT_MODELS if m in row_map]

    acc_row = ["Final5 TestAcc"] + [_fmt(row_map[m]["Final5_TestAcc_mean"], row_map[m]["Final5_TestAcc_std"], 4, 4) for m in models]
    time_row = ["Time (mins)"] + [_fmt(row_map[m]["TotalTime(min)_mean"], row_map[m]["TotalTime(min)_std"], 2, 3) for m in models]
    param_row = ["Final ParamCount"] + [_fmt(row_map[m]["FinalParamCount_mean"], row_map[m]["FinalParamCount_std"], 0, 0) for m in models]

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    for j, v in enumerate(header, start=1):
        ws.cell(row=1, column=j, value=v)
    for j, v in enumerate(acc_row, start=1):
        ws.cell(row=2, column=j, value=v)
    for j, v in enumerate(time_row, start=1):
        ws.cell(row=3, column=j, value=v)
    for j, v in enumerate(param_row, start=1):
        ws.cell(row=4, column=j, value=v)

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font = bold
        cell.alignment = center
        cell.border = border

    for r in [2, 3, 4]:
        ws.cell(row=r, column=1).alignment = left
        ws.cell(row=r, column=1).border = border
        for c in range(2, len(header) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = center
            cell.border = border

    ws.column_dimensions["A"].width = 20
    for idx in range(2, len(header) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 18

    wb.save(out_xlsx_path)


# =========================
# 绘图
# =========================
def _iter_plot_models(agg: Dict[str, dict]) -> List[str]:
    return [m for m in DEFAULT_MODELS if m in agg]


def plot_metric_vs_time_mean_std(
    agg: Dict[str, dict],
    out_png: str,
    ylabel: str,
    title: str,
    ylim: Optional[Tuple[float, float]] = None,
) -> None:
    plt.figure()
    for model in _iter_plot_models(agg):
        a = agg[model]
        x = a["mean_time"]
        y = a["mean_value"]
        s = a["std_value"]
        mask = np.isfinite(x) & np.isfinite(y) & np.isfinite(s)
        x, y, s = x[mask], y[mask], s[mask]
        if x.size == 0:
            continue
        plt.plot(x, y, label=DISPLAY_NAME_MAP.get(model, model), color=COLOR_MAP.get(model, None))
        plt.fill_between(x, y - s, y + s, color=COLOR_MAP.get(model, None), alpha=0.2)

    plt.xlabel("Time (s)")
    plt.ylabel(ylabel)
    plt.title(title)
    _apply_ylim(ylim)
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    _ensure_dir(os.path.dirname(out_png))
    plt.savefig(out_png, dpi=160)
    plt.close()


def plot_metric_vs_epoch_mean_std(
    agg: Dict[str, dict],
    out_png: str,
    ylabel: str,
    title: str,
    ylim: Optional[Tuple[float, float]] = None,
) -> None:
    plt.figure()
    for model in _iter_plot_models(agg):
        a = agg[model]
        x = a["epochs"].astype(np.float64)
        y = a["mean_value"]
        s = a["std_value"]
        mask = np.isfinite(x) & np.isfinite(y) & np.isfinite(s)
        x, y, s = x[mask], y[mask], s[mask]
        if x.size == 0:
            continue
        plt.plot(x, y, label=DISPLAY_NAME_MAP.get(model, model), color=COLOR_MAP.get(model, None))
        plt.fill_between(x, y - s, y + s, color=COLOR_MAP.get(model, None), alpha=0.2)

    plt.xlabel("Epoch")
    plt.ylabel(ylabel)
    plt.title(title)
    _apply_ylim(ylim)
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    _ensure_dir(os.path.dirname(out_png))
    plt.savefig(out_png, dpi=160)
    plt.close()


def plot_final_acc_bar(
    summary_rows: List[dict],
    out_png: str,
    ylim: Optional[Tuple[float, float]] = None,
) -> None:
    row_map = {r["Model"]: r for r in summary_rows}
    models = [m for m in DEFAULT_MODELS if m in row_map]
    xs = np.arange(len(models), dtype=np.float64)
    ys = np.array([row_map[m]["Final5_TestAcc_mean"] for m in models], dtype=np.float64)
    es = np.array([row_map[m]["Final5_TestAcc_std"] for m in models], dtype=np.float64)
    es = np.where(np.isfinite(es), es, 0.0)

    plt.figure()
    plt.bar(xs, ys, color=[COLOR_MAP.get(m, None) for m in models])
    plt.errorbar(xs, ys, yerr=es, fmt="none", capsize=8, capthick=1.8, elinewidth=1.8, ecolor="black")
    plt.xticks(xs, [DISPLAY_NAME_MAP.get(m, m) for m in models])
    plt.ylabel("Final5 TestAcc")
    plt.title("Final Test Accuracy (mean ± std over 3 runs)")
    _apply_ylim(ylim)
    plt.grid(True, axis="y")
    plt.tight_layout()
    _ensure_dir(os.path.dirname(out_png))
    plt.savefig(out_png, dpi=160)
    plt.close()


# =========================
# 主函数
# =========================
def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Aggregate and plot current 4-mode VGG results.")
    parser.add_argument(
        "--inputs",
        nargs="+",
        default=DEFAULT_INPUT_FILES,
        help="results_run_*.txt 文件路径列表",
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        help="输出目录；不写则自动放到输入文件公共目录下的 summary_four_modes/",
    )
    parser.add_argument(
        "--last-k",
        type=int,
        default=FINAL_LAST_K,
        help="最终精度 summary 里取最后多少个 epoch 求平均",
    )
    parser.add_argument(
        "--epoch-union",
        action="store_true",
        help="默认使用 epoch intersection；加这个参数则改为 epoch union",
    )
    parser.add_argument("--ylim-test-acc-time", nargs=2, type=float, default=None, metavar=("YMIN", "YMAX"))
    parser.add_argument("--ylim-test-acc-epoch", nargs=2, type=float, default=None, metavar=("YMIN", "YMAX"))
    parser.add_argument("--ylim-train-loss-time", nargs=2, type=float, default=None, metavar=("YMIN", "YMAX"))
    parser.add_argument("--ylim-train-acc-time", nargs=2, type=float, default=None, metavar=("YMIN", "YMAX"))
    parser.add_argument("--ylim-test-loss-time", nargs=2, type=float, default=None, metavar=("YMIN", "YMAX"))
    parser.add_argument("--ylim-param-count-epoch", nargs=2, type=float, default=None, metavar=("YMIN", "YMAX"))
    parser.add_argument("--ylim-final-acc-bar", nargs=2, type=float, default=None, metavar=("YMIN", "YMAX"))
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    for p in args.inputs:
        if not os.path.isfile(p):
            raise FileNotFoundError(f"找不到输入文件：{p}")

    output_dir = args.output_dir or _default_output_dir(args.inputs)
    _ensure_dir(output_dir)

    run_dicts = [read_results_tsv_allmetrics(p) for p in args.inputs]
    use_intersection = (not args.epoch_union)

    ylim_test_acc_time = _resolve_ylim(_parse_ylim_arg(args.ylim_test_acc_time), YLIM_TEST_ACC_TIME)
    ylim_test_acc_epoch = _resolve_ylim(_parse_ylim_arg(args.ylim_test_acc_epoch), YLIM_TEST_ACC_EPOCH)
    ylim_train_loss_time = _resolve_ylim(_parse_ylim_arg(args.ylim_train_loss_time), YLIM_TRAIN_LOSS_TIME)
    ylim_train_acc_time = _resolve_ylim(_parse_ylim_arg(args.ylim_train_acc_time), YLIM_TRAIN_ACC_TIME)
    ylim_test_loss_time = _resolve_ylim(_parse_ylim_arg(args.ylim_test_loss_time), YLIM_TEST_LOSS_TIME)
    ylim_param_count_epoch = _resolve_ylim(_parse_ylim_arg(args.ylim_param_count_epoch), YLIM_PARAM_COUNT_EPOCH)
    ylim_final_acc_bar = _resolve_ylim(_parse_ylim_arg(args.ylim_final_acc_bar), YLIM_FINAL_ACC_BAR)

    # 逐指标聚合
    agg_test_acc = aggregate_runs_metric(run_dicts, "test_acc", models=DEFAULT_MODELS, use_epoch_intersection=use_intersection)
    agg_train_loss = aggregate_runs_metric(run_dicts, "train_loss", models=DEFAULT_MODELS, use_epoch_intersection=use_intersection)
    agg_train_acc = aggregate_runs_metric(run_dicts, "train_acc", models=DEFAULT_MODELS, use_epoch_intersection=use_intersection)
    agg_test_loss = aggregate_runs_metric(run_dicts, "test_loss", models=DEFAULT_MODELS, use_epoch_intersection=use_intersection)
    agg_param_count = aggregate_runs_metric(run_dicts, "param_count", models=DEFAULT_MODELS, use_epoch_intersection=use_intersection)

    if not agg_test_acc:
        raise ValueError("没有可用的 test_acc 数据，请检查输入文件。")

    # 保存按 epoch 聚合后的 tsv
    save_agg_tsv(agg_test_acc, os.path.join(output_dir, "test_accuracy_vs_time_mean.tsv"), "TestAcc")
    save_agg_tsv(agg_train_loss, os.path.join(output_dir, "train_loss_vs_time_mean.tsv"), "TrainLoss")
    save_agg_tsv(agg_train_acc, os.path.join(output_dir, "train_accuracy_vs_time_mean.tsv"), "TrainAcc")
    save_agg_tsv(agg_test_loss, os.path.join(output_dir, "test_loss_vs_time_mean.tsv"), "TestLoss")
    save_agg_tsv(agg_param_count, os.path.join(output_dir, "param_count_vs_time_mean.tsv"), "ParamCount")

    # 曲线图
    plot_metric_vs_time_mean_std(
        agg_test_acc,
        os.path.join(output_dir, "test_accuracy_vs_time_mean.png"),
        ylabel="Test Accuracy",
        title="Test Accuracy vs Time (mean ± std over runs)",
        ylim=ylim_test_acc_time,
    )
    plot_metric_vs_epoch_mean_std(
        agg_test_acc,
        os.path.join(output_dir, "test_accuracy_vs_epoch_mean.png"),
        ylabel="Test Accuracy",
        title="Test Accuracy vs Epoch (mean ± std over runs)",
        ylim=ylim_test_acc_epoch,
    )
    plot_metric_vs_time_mean_std(
        agg_train_loss,
        os.path.join(output_dir, "train_loss_vs_time_mean.png"),
        ylabel="Train Loss",
        title="Train Loss vs Time (mean ± std over runs)",
        ylim=ylim_train_loss_time,
    )
    plot_metric_vs_time_mean_std(
        agg_train_acc,
        os.path.join(output_dir, "train_accuracy_vs_time_mean.png"),
        ylabel="Train Accuracy",
        title="Train Accuracy vs Time (mean ± std over runs)",
        ylim=ylim_train_acc_time,
    )
    plot_metric_vs_time_mean_std(
        agg_test_loss,
        os.path.join(output_dir, "test_loss_vs_time_mean.png"),
        ylabel="Test Loss",
        title="Test Loss vs Time (mean ± std over runs)",
        ylim=ylim_test_loss_time,
    )
    plot_metric_vs_epoch_mean_std(
        agg_param_count,
        os.path.join(output_dir, "param_count_vs_epoch_mean.png"),
        ylabel="Parameter Count",
        title="Parameter Count vs Epoch (mean ± std over runs)",
        ylim=ylim_param_count_epoch,
    )

    # 最终 summary
    per_run_rows, summary_rows = compute_final_metrics(run_dicts, models=DEFAULT_MODELS, last_k=args.last_k)
    save_final_metrics_csv(per_run_rows, summary_rows, output_dir)
    save_final_metrics_excel(summary_rows, os.path.join(output_dir, "final_metrics_summary.xlsx"))
    plot_final_acc_bar(summary_rows, os.path.join(output_dir, "final_testacc_bar_mean_std.png"), ylim=ylim_final_acc_bar)

    print("已保存到：", output_dir)
    print("  - test_accuracy_vs_time_mean.png")
    print("  - test_accuracy_vs_epoch_mean.png")
    print("  - train_loss_vs_time_mean.png")
    print("  - train_accuracy_vs_time_mean.png")
    print("  - test_loss_vs_time_mean.png")
    print("  - param_count_vs_epoch_mean.png")
    print("  - final_metrics_per_run.csv")
    print("  - final_metrics_summary.csv")
    print("  - final_metrics_summary.xlsx（若环境有 openpyxl）")
    print("  - final_testacc_bar_mean_std.png")


if __name__ == "__main__":
    main()